#!/usr/bin/env python3
"""
Pharmacy Quotation Automation System v3
========================================
- Supports any PDF format (text or scanned)
- Includes VAT tracking from historical data
- Generates clean output for sending to companies
- Better item extraction with multiple parsing strategies
"""

import pandas as pd
import pdfplumber
import re
import os
from datetime import datetime
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False


class PriceDatabase:
    """Price database with VAT tracking"""
    
    def __init__(self):
        self.items = {}
        self.row_counter = 0
    
    def add_item(self, item_name, price, vat_rate=0, original_name=None):
        """Add item with price and VAT rate"""
        if not item_name or price is None:
            return
        
        normalized = self._normalize(item_name)
        if len(normalized) < 3:
            return
        
        self.row_counter += 1
        self.items[normalized] = {
            'price': float(price),
            'vat_rate': float(vat_rate) if vat_rate else 0,
            'original_name': original_name or item_name,
            'last_used': self.row_counter
        }
    
    def _normalize(self, name):
        if not name:
            return ""
        name = str(name).upper().strip()
        name = re.sub(r'\s+', ' ', name)
        name = name.replace('\n', ' ')
        return name
    
    def find_price(self, item_name, threshold=55):
        """Find price with VAT info"""
        search_name = self._normalize(item_name)
        
        # Exact match
        if search_name in self.items:
            item = self.items[search_name]
            return {
                'found': True,
                'price': item['price'],
                'vat_rate': item['vat_rate'],
                'matched_to': item['original_name'],
                'confidence': 100,
                'status': 'found'
            }
        
        # Fuzzy matching
        all_names = list(self.items.keys())
        if not all_names:
            return self._not_found()
        
        best_match = None
        best_score = 0
        
        for scorer in [fuzz.token_set_ratio, fuzz.token_sort_ratio, fuzz.partial_ratio]:
            match = process.extractOne(search_name, all_names, scorer=scorer)
            if match and match[1] > best_score:
                len_ratio = min(len(search_name), len(match[0])) / max(len(search_name), len(match[0]))
                if len_ratio > 0.15:
                    best_match = match
                    best_score = match[1]
        
        if best_match and best_score >= threshold:
            item = self.items[best_match[0]]
            status = 'found' if best_score >= 75 else 'review'
            return {
                'found': True,
                'price': item['price'],
                'vat_rate': item['vat_rate'],
                'matched_to': item['original_name'],
                'confidence': best_score,
                'status': status
            }
        
        return self._not_found(best_score)
    
    def _not_found(self, score=0):
        return {
            'found': False,
            'price': None,
            'vat_rate': 0.05,  # Default 5% VAT for unknown items
            'matched_to': None,
            'confidence': score,
            'status': 'notfound'
        }


class RequestParser:
    """Parses request files (PDF or Excel)"""
    
    @staticmethod
    def extract_items(file_path):
        """Extract items from PDF or Excel request file"""
        ext = file_path.lower().split('.')[-1]
        
        if ext in ['xlsx', 'xls']:
            return RequestParser._parse_excel(file_path)
        elif ext == 'pdf':
            return PDFParser.extract_items(file_path)
        else:
            print(f"  → Unsupported file type: {ext}")
            return []
    
    @staticmethod
    def _parse_excel(file_path):
        """Parse Excel request file"""
        items = []
        
        try:
            all_sheets = pd.read_excel(file_path, sheet_name=None, header=None)
        except Exception as e:
            print(f"  → Error reading Excel: {e}")
            return items
        
        for sheet_name, df in all_sheets.items():
            sheet_items = RequestParser._extract_items_from_sheet(df)
            items.extend(sheet_items)
            if sheet_items:
                print(f"  → Found {len(sheet_items)} items in sheet '{sheet_name}'")
        
        # Remove duplicates based on description
        seen = set()
        unique_items = []
        for item in items:
            key = item['description'].upper()
            if key not in seen:
                seen.add(key)
                unique_items.append(item)
        
        return unique_items
    
    @staticmethod
    def _extract_items_from_sheet(df):
        """Extract items from a single sheet"""
        items = []
        
        # Try to detect header row and column structure
        # Common patterns:
        # - [SL NO, ITEM/DESCRIPTION/MATERIAL NAME, UNIT, QTY, ...]
        # - [No, Description, Unit, Quantity, Price, ...]
        
        header_keywords = ['item', 'description', 'material', 'name', 'product']
        qty_keywords = ['qty', 'quantity', 'qnty']
        unit_keywords = ['unit', 'uom', 'u/m']
        
        desc_col = None
        qty_col = None
        unit_col = None
        header_row = None
        
        # Find header row
        for idx, row in df.iterrows():
            row_str = ' '.join([str(v).lower() for v in row.tolist() if pd.notna(v)])
            if any(kw in row_str for kw in header_keywords):
                header_row = idx
                # Find column indices
                for col_idx, val in enumerate(row.tolist()):
                    if pd.isna(val):
                        continue
                    val_lower = str(val).lower()
                    if any(kw in val_lower for kw in header_keywords) and desc_col is None:
                        desc_col = col_idx
                    elif any(kw in val_lower for kw in qty_keywords) and qty_col is None:
                        qty_col = col_idx
                    elif any(kw in val_lower for kw in unit_keywords) and unit_col is None:
                        unit_col = col_idx
                break
        
        # If no header found, try common column positions
        if desc_col is None:
            # Try column 1 for description (column 0 often has serial number)
            desc_col = 1
            qty_col = 3
            unit_col = 2
            header_row = -1
        
        # Extract items
        for idx, row in df.iterrows():
            if header_row is not None and idx <= header_row:
                continue
            
            vals = row.tolist()
            
            # Get description
            if desc_col >= len(vals):
                continue
            description = vals[desc_col]
            if pd.isna(description) or len(str(description).strip()) < 3:
                continue
            description = str(description).strip()
            
            # Skip if it looks like a header or total
            if description.upper() in ['TOTAL', 'SUBTOTAL', 'GRAND TOTAL', 'NAN', 'DESCRIPTION', 'ITEM']:
                continue
            
            # Get quantity
            qty = None
            if qty_col is not None and qty_col < len(vals):
                try:
                    qty = int(float(vals[qty_col]))
                except:
                    pass
            
            # If no qty found, search all columns for a number
            if qty is None:
                for v in vals:
                    if pd.notna(v):
                        try:
                            potential_qty = int(float(v))
                            if 0 < potential_qty < 10000:
                                qty = potential_qty
                                break
                        except:
                            pass
            
            if qty is None or qty <= 0:
                continue
            
            # Get unit
            unit = 'PCS'
            if unit_col is not None and unit_col < len(vals) and pd.notna(vals[unit_col]):
                unit = str(vals[unit_col]).strip().upper()
            
            items.append({
                'description': description,
                'quantity': qty,
                'unit': unit
            })
        
        return items


class PDFParser:
    """Universal PDF parser - works with any format"""
    
    @staticmethod
    def extract_items(pdf_path):
        """Extract items from any PDF format"""
        # Try text extraction first
        all_text = PDFParser._extract_text(pdf_path)
        
        # If no text, try OCR
        if not all_text.strip() and OCR_AVAILABLE:
            print("  → Using OCR for scanned PDF...")
            all_text = PDFParser._ocr_pdf(pdf_path)
        
        if not all_text.strip():
            return []
        
        # Try multiple parsing strategies and combine results
        items = []
        
        # Strategy 1: Numbered list (like First Aid PDF)
        items1 = PDFParser._parse_numbered_list(all_text)
        if items1:
            items.extend(items1)
        
        # Strategy 2: Table format with pipes
        items2 = PDFParser._parse_table_format(all_text)
        for item in items2:
            # Avoid duplicates
            if not any(i['description'].upper() == item['description'].upper() for i in items):
                items.append(item)
        
        # Strategy 3: Description + Quantity patterns (no leading number)
        items3 = PDFParser._parse_description_qty_patterns(all_text)
        for item in items3:
            if not any(i['description'].upper() == item['description'].upper() for i in items):
                items.append(item)
        
        return items
    
    @staticmethod
    def _parse_description_qty_patterns(text):
        """Parse lines with Description followed by quantity"""
        items = []
        lines = text.split('\n')
        
        unit_words = ['pcs', 'pkts', 'pkt', 'box', 'boxes', 'nos', 'rolls', 'each', 'unit', 
                      'pieces', 'pack', 'bottles', 'bottle', 'tubes', 'carton', 'tins', 'ea', 
                      'pc', 'litre', 'pks', 'pairs', 'dzn', 'doz', 'dozen']
        
        for line in lines:
            line = line.strip()
            if not line or len(line) < 5:
                continue
            
            # Skip obvious headers
            upper = line.upper()
            if any(skip in upper for skip in ['DESCRIPTION', 'QUANTITY', 'TOTAL', 'SL.NO', 
                                               'CHECK LIST', 'INVENTORY', 'ITEMS', 'MATERIALS']):
                continue
            
            # Pattern: "Item Name 10PCS" or "Item Name 10 PCS"
            match = re.search(r'^(.+?)\s+(\d+)\s*(' + '|'.join(unit_words) + r')s?\s*$', line, re.IGNORECASE)
            if match:
                description = match.group(1).strip()
                qty = int(match.group(2))
                unit = match.group(3).upper()
                
                # Clean description - remove leading numbers
                description = re.sub(r'^\d+[\s.]+', '', description)
                
                if description and len(description) > 2 and 0 < qty < 50000:
                    items.append({
                        'description': description,
                        'quantity': qty,
                        'unit': unit
                    })
                continue
            
            # Pattern: "Item Name 10PKTS" (no space)
            match = re.search(r'^(.+?)\s*(\d+)(' + '|'.join(unit_words) + r')s?\s*$', line, re.IGNORECASE)
            if match:
                description = match.group(1).strip()
                qty = int(match.group(2))
                unit = match.group(3).upper()
                
                description = re.sub(r'^\d+[\s.]+', '', description)
                
                if description and len(description) > 2 and 0 < qty < 50000:
                    items.append({
                        'description': description,
                        'quantity': qty,
                        'unit': unit
                    })
        
        return items
    
    @staticmethod
    def _extract_text(pdf_path):
        """Extract text from PDF"""
        all_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # Try table extraction first
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row:
                            all_text += ' | '.join([str(c) if c else '' for c in row]) + '\n'
                
                # Also get regular text
                text = page.extract_text()
                if text:
                    all_text += text + '\n'
        return all_text
    
    @staticmethod
    def _ocr_pdf(pdf_path):
        """OCR for scanned PDFs with optimized settings"""
        images = convert_from_path(pdf_path, dpi=400)
        all_text = ""
        
        # Try multiple PSM modes and combine results
        psm_modes = [3, 12]  # Auto and sparse text
        
        for image in images:
            page_text = ""
            for psm in psm_modes:
                config = f'--oem 3 --psm {psm}'
                text = pytesseract.image_to_string(image, config=config)
                page_text += text + "\n"
            all_text += page_text + "\n"
        
        return all_text
    
    @staticmethod
    def _parse_numbered_list(text):
        """Parse numbered list format (like the First Aid PDF)"""
        items = []
        lines = text.split('\n')
        
        # Unit patterns - expanded
        unit_words = ['pcs', 'pkts', 'pkt', 'box', 'nos', 'rolls', 'each', 'unit', 
                      'pieces', 'pack', 'bottles', 'tubes', 'carton', 'tins', 'ea', 
                      'pc', 'litre', 'pks', 'pairs', 'pes', 'size']
        unit_pattern = r'(\d+)\s*(' + '|'.join(unit_words) + r')\.?\s*$'
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            
            # Skip headers
            upper = line.upper()
            if any(skip in upper for skip in ['TOTAL', 'S.NO', 'ITEMS', 'QUANTITY', 
                                               'FORM REF', 'REQUIREMENTS', 'DESCRIPTION']):
                i += 1
                continue
            
            # Look for pattern: "NUMBER DESCRIPTION QUANTITY"
            match = re.match(r'^(\d{1,3})[\s.|]+(.+)', line)
            if match:
                item_num = int(match.group(1))
                rest = match.group(2).strip()
                
                # Build full text including continuation lines
                full_text = rest
                j = i + 1
                while j < len(lines) and j < i + 4:
                    next_line = lines[j].strip()
                    if re.match(r'^\d{1,3}[\s.|]', next_line):
                        break
                    if not next_line or any(skip in next_line.upper() for skip in ['TOTAL', 'FORM REF']):
                        break
                    full_text += ' ' + next_line
                    j += 1
                
                # Try multiple extraction strategies
                qty = None
                unit = 'PCS'
                description = None
                
                # Strategy 1: Unit word at end
                qty_match = re.search(unit_pattern, full_text, re.IGNORECASE)
                if qty_match:
                    qty = int(qty_match.group(1))
                    unit = qty_match.group(2).upper()
                    description = full_text[:qty_match.start()].strip()
                
                # Strategy 2: Just number at end (no unit word)
                if not qty:
                    qty_match = re.search(r'\s(\d{1,5})\s*$', full_text)
                    if qty_match:
                        qty = int(qty_match.group(1))
                        description = full_text[:qty_match.start()].strip()
                
                # Strategy 3: Number with "each size" or similar
                if not qty:
                    qty_match = re.search(r'(\d+)\s*(?:pkts?\s+)?each\s+size', full_text, re.IGNORECASE)
                    if qty_match:
                        qty = int(qty_match.group(1))
                        description = full_text[:qty_match.start()].strip()
                
                # Strategy 4: Embedded quantity like "20pcs" or "10 pcs" anywhere
                if not qty:
                    qty_match = re.search(r'(\d+)\s*(pcs|pkts|box|rolls|pieces)', full_text, re.IGNORECASE)
                    if qty_match:
                        qty = int(qty_match.group(1))
                        unit = qty_match.group(2).upper()
                        # Description is text before qty
                        description = full_text[:qty_match.start()].strip()
                        if not description:
                            description = full_text[qty_match.end():].strip()
                
                # Strategy 5: Just take the line with a default qty of 1 if nothing else works
                if not qty and len(full_text) > 5:
                    # Check if there's ANY number we can use
                    all_nums = re.findall(r'\b(\d{1,4})\b', full_text)
                    if all_nums:
                        # Use last number as quantity
                        qty = int(all_nums[-1])
                        description = re.sub(r'\s*\d{1,4}\s*$', '', full_text).strip()
                
                # Clean description
                if description:
                    description = re.sub(r'[|]', '', description)
                    description = re.sub(r'\s+', ' ', description).strip()
                    description = re.sub(r'^[\s.,|]+', '', description)
                    description = re.sub(r'[\s.,|]+$', '', description)
                
                if description and len(description) > 2 and qty and 0 < qty < 50000:
                    items.append({
                        'description': description,
                        'quantity': qty,
                        'unit': unit
                    })
            
            i += 1
        
        return items
    
    @staticmethod
    def _parse_table_format(text):
        """Parse table format (like Tornado PDF or inventory lists)"""
        items = []
        lines = text.split('\n')
        
        # Unit patterns
        unit_words = ['pcs', 'pkts', 'pkt', 'box', 'boxes', 'nos', 'rolls', 'each', 'unit', 
                      'pieces', 'pack', 'bottles', 'bottle', 'tubes', 'carton', 'tins', 'ea', 
                      'pc', 'litre', 'pks', 'pairs', 'dzn', 'doz', 'dozen']
        
        for line in lines:
            line = line.strip()
            if not line or len(line) < 5:
                continue
            
            # Skip headers
            upper = line.upper()
            if any(skip in upper for skip in ['DESCRIPTION', 'QUANTITY', 'TOTAL', 'SL.NO', 
                                               'CHECK LIST', 'INVENTORY', 'ITEMS']):
                continue
            
            # Strategy 1: Parse pipe-separated tables
            if '|' in line:
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if len(parts) >= 3:
                    for i, part in enumerate(parts):
                        if len(part) > 5 and not part.replace('.', '').replace('-', '').isdigit():
                            # Potential description - look for quantity
                            for j in range(i+1, min(i+3, len(parts))):
                                try:
                                    qty = int(parts[j])
                                    if 0 < qty < 50000:
                                        unit = parts[j+1] if j+1 < len(parts) else 'PCS'
                                        items.append({
                                            'description': part,
                                            'quantity': qty,
                                            'unit': unit.upper()
                                        })
                                        break
                                except:
                                    continue
                            break
            
            # Strategy 2: Look for "Description QTY+UNIT" pattern (like the inventory list)
            # Pattern: "Plastic Band Aids 12PKTS" or "Deep Heat Spray 10PCS"
            qty_unit_match = re.search(r'(\d+)\s*(' + '|'.join(unit_words) + r')\s*$', line, re.IGNORECASE)
            if qty_unit_match:
                qty = int(qty_unit_match.group(1))
                unit = qty_unit_match.group(2).upper()
                description = line[:qty_unit_match.start()].strip()
                
                # Remove leading numbers (item numbers)
                description = re.sub(r'^\d+\s+', '', description)
                description = re.sub(r'^[\s|]+', '', description)
                
                if description and len(description) > 2 and 0 < qty < 50000:
                    items.append({
                        'description': description,
                        'quantity': qty,
                        'unit': unit
                    })
                continue
            
            # Strategy 3: Look for "QTY UNIT" at end (like "5 Bottle" or "1pcs")
            qty_match = re.search(r'(\d+)\s+(' + '|'.join(unit_words) + r')\s*$', line, re.IGNORECASE)
            if qty_match:
                qty = int(qty_match.group(1))
                unit = qty_match.group(2).upper()
                description = line[:qty_match.start()].strip()
                description = re.sub(r'^\d+\s+', '', description)
                
                if description and len(description) > 2 and 0 < qty < 50000:
                    items.append({
                        'description': description,
                        'quantity': qty,
                        'unit': unit
                    })
        
        return items


class ExcelLoader:
    """Load prices with VAT from Excel - smart column detection"""
    
    @staticmethod
    def load_prices(excel_path, price_db):
        """Load prices and VAT from all sheets"""
        try:
            all_sheets = pd.read_excel(excel_path, sheet_name=None, header=None)
        except Exception as e:
            print(f"  → Error: {e}")
            return 0
        
        total_loaded = 0
        
        for sheet_name, df in all_sheets.items():
            loaded = ExcelLoader._process_sheet(df, price_db, sheet_name)
            total_loaded += loaded
        
        return total_loaded
    
    @staticmethod
    def _process_sheet(df, price_db, sheet_name):
        """Process sheet with smart column detection"""
        loaded = 0
        
        for idx, row in df.iterrows():
            vals = row.tolist()
            if len(vals) < 5:
                continue
            
            # Find description column (first column with valid text)
            desc_col = None
            item_name = None
            
            for i, val in enumerate(vals[:4]):
                if ExcelLoader._is_valid_name(val):
                    desc_col = i
                    item_name = str(val).strip()
                    break
            
            if item_name is None:
                continue
            
            # Now find the unit price using the formula: QTY * PRICE ≈ TOTAL
            # Try different column combinations
            price = None
            vat_amount = 0
            
            # Strategy: Find columns where col_a * col_b ≈ col_c (within 1%)
            numeric_cols = []
            for i in range(desc_col + 1, min(len(vals), desc_col + 8)):
                try:
                    num = float(vals[i])
                    if pd.notna(num) and num > 0:
                        numeric_cols.append((i, num))
                except:
                    pass
            
            # Try to find QTY, PRICE, TOTAL pattern
            for i, (qty_col, qty_val) in enumerate(numeric_cols):
                for j, (price_col, price_val) in enumerate(numeric_cols):
                    if i >= j:
                        continue
                    expected_total = qty_val * price_val
                    
                    # Look for a total column
                    for k, (total_col, total_val) in enumerate(numeric_cols):
                        if k <= j:
                            continue
                        
                        # Check if qty * price ≈ total (within 5% or 1 AED)
                        diff = abs(expected_total - total_val)
                        if diff < 1 or (expected_total > 0 and diff / expected_total < 0.05):
                            # Found the pattern! price_val is the unit price
                            price = price_val
                            
                            # Look for VAT in next columns
                            for m, (vat_col, vat_val) in enumerate(numeric_cols):
                                if vat_col > total_col and vat_val < total_val and vat_val > 0:
                                    # Check if this could be VAT (typically 5% of total)
                                    if abs(vat_val - total_val * 0.05) < total_val * 0.02:
                                        vat_amount = vat_val
                                        break
                            break
                    if price:
                        break
                if price:
                    break
            
            # Fallback: if no pattern found, try column 4 (common position for unit price)
            if price is None and len(vals) > 4:
                try:
                    potential_price = float(vals[4]) if desc_col < 2 else float(vals[desc_col + 3])
                    if 0.1 < potential_price < 10000:
                        price = potential_price
                except:
                    pass
            
            if item_name and price and price > 0:
                # Calculate VAT rate
                vat_rate = 0
                if vat_amount > 0 and price > 0:
                    # VAT rate based on unit price
                    vat_rate = 0.05  # Default 5% if VAT exists
                
                price_db.add_item(item_name, price, vat_rate, item_name)
                loaded += 1
        
        return loaded
    
    @staticmethod
    def _is_valid_name(val):
        if pd.isna(val):
            return False
        s = str(val).strip()
        if len(s) < 4:
            return False
        # Skip common headers and keywords
        skip_words = ['TOTAL', 'TOTAL=', 'NAN', 'ITEM', 'DESCRIPTION', 
                      'ITEM DESCRIPTION', 'SN', 'SN ', 'AMOUNT', 'VAT',
                      'SLNO', 'NET VALUE', 'T PRICE', 'U PRICE', 'MATERIAL NAME',
                      'UOM', 'QTY', 'UNIT', 'QUANTITY', 'PRICE']
        if s.upper() in skip_words:
            return False
        if re.match(r'^[\d\s.,]+$', s):
            return False
        # Must contain at least some letters
        if not re.search(r'[a-zA-Z]{2,}', s):
            return False
        return True


class QuotationGenerator:
    """Generate Excel with internal and clean sheets"""
    
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.colors = {
            'header': '1F4E79',
            'found': 'C6EFCE',
            'review': 'FFEB9C',
            'notfound': 'FFC7CE',
        }
    
    def generate(self, items, price_db, output_path, title=""):
        """Generate quotation with multiple sheets"""
        wb = Workbook()
        
        # Sheet 1: Clean quotation for company (no confidence)
        ws_clean = wb.active
        ws_clean.title = "Quotation"
        results = self._create_clean_sheet(ws_clean, items, price_db, title)
        
        # Sheet 2: Internal with confidence scores
        ws_internal = wb.create_sheet("Internal Review")
        self._create_internal_sheet(ws_internal, items, price_db, results, title)
        
        # Sheet 3: Items not found
        ws_notfound = wb.create_sheet("Items to Price")
        self._create_notfound_sheet(ws_notfound, items, results)
        
        wb.save(output_path)
        
        stats = {
            'found': sum(1 for r in results if r['status'] == 'found'),
            'review': sum(1 for r in results if r['status'] == 'review'),
            'notfound': sum(1 for r in results if r['status'] == 'notfound')
        }
        return stats
    
    def _create_clean_sheet(self, ws, items, price_db, title):
        """Clean quotation for sending to company"""
        results = []
        
        # Header
        ws['A1'] = f"QUOTATION"
        ws['A1'].font = Font(bold=True, size=18, color=self.colors['header'])
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
        if title:
            ws['A3'] = f"Reference: {title}"
        
        # Column headers
        headers = ['No.', 'Description', 'Unit', 'Qty', 'Unit Price', 'Amount', 'VAT', 'Gross Amount']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=self.colors['header'])
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        # Data
        row = 6
        for i, item in enumerate(items, 1):
            match = price_db.find_price(item['description'])
            results.append({**item, **match})
            
            ws.cell(row=row, column=1, value=i).border = self.thin_border
            ws.cell(row=row, column=2, value=item['description']).border = self.thin_border
            ws.cell(row=row, column=3, value=item['unit']).border = self.thin_border
            ws.cell(row=row, column=4, value=item['quantity']).border = self.thin_border
            
            # Price
            price_cell = ws.cell(row=row, column=5)
            price_cell.border = self.thin_border
            price_cell.number_format = '#,##0.00'
            
            # Amount
            amount_cell = ws.cell(row=row, column=6)
            amount_cell.border = self.thin_border
            amount_cell.number_format = '#,##0.00'
            
            # VAT
            vat_cell = ws.cell(row=row, column=7)
            vat_cell.border = self.thin_border
            vat_cell.number_format = '#,##0.00'
            
            # Gross
            gross_cell = ws.cell(row=row, column=8)
            gross_cell.border = self.thin_border
            gross_cell.number_format = '#,##0.00'
            
            if match['found']:
                price_cell.value = match['price']
                amount_cell.value = f'=D{row}*E{row}'
                vat_pct = match['vat_rate'] if match['vat_rate'] > 0 else 0.05
                vat_cell.value = f'=F{row}*{vat_pct}'
                gross_cell.value = f'=F{row}+G{row}'
            else:
                # Leave blank for manual entry
                amount_cell.value = f'=IFERROR(D{row}*E{row},"")'
                vat_cell.value = f'=IFERROR(F{row}*0.05,"")'
                gross_cell.value = f'=IFERROR(F{row}+G{row},"")'
            
            row += 1
        
        # Totals
        row += 1
        ws.cell(row=row, column=5, value="TOTALS:").font = Font(bold=True)
        
        for col, label in [(6, 'Amount'), (7, 'VAT'), (8, 'Gross')]:
            cell = ws.cell(row=row, column=col, value=f'=SUM({chr(64+col)}6:{chr(64+col)}{row-2})')
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            cell.border = self.thin_border
            if col == 8:
                cell.fill = PatternFill('solid', fgColor=self.colors['header'])
                cell.font = Font(bold=True, color='FFFFFF')
        
        # Column widths
        widths = [6, 50, 10, 8, 14, 14, 12, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i)].width = w
        
        return results
    
    def _create_internal_sheet(self, ws, items, price_db, results, title):
        """Internal sheet with confidence scores"""
        ws['A1'] = "INTERNAL REVIEW - DO NOT SEND TO CUSTOMER"
        ws['A1'].font = Font(bold=True, size=14, color='FF0000')
        ws.merge_cells('A1:I1')
        
        headers = ['No.', 'Description', 'Qty', 'Unit Price', 'Amount', 'VAT Rate', 
                   'Matched To', 'Confidence', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=self.colors['header'])
            cell.border = self.thin_border
        
        row = 4
        for i, r in enumerate(results, 1):
            ws.cell(row=row, column=1, value=i).border = self.thin_border
            ws.cell(row=row, column=2, value=r['description']).border = self.thin_border
            ws.cell(row=row, column=3, value=r['quantity']).border = self.thin_border
            
            price_cell = ws.cell(row=row, column=4, value=r.get('price', ''))
            price_cell.border = self.thin_border
            price_cell.number_format = '#,##0.00'
            
            if r.get('price'):
                ws.cell(row=row, column=5, value=r['price'] * r['quantity']).border = self.thin_border
            else:
                ws.cell(row=row, column=5, value='').border = self.thin_border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            vat_rate = r.get('vat_rate', 0)
            ws.cell(row=row, column=6, value=f"{vat_rate*100:.1f}%" if vat_rate else "5.0%").border = self.thin_border
            
            ws.cell(row=row, column=7, value=r.get('matched_to', '')).border = self.thin_border
            ws.cell(row=row, column=8, value=f"{r['confidence']}%").border = self.thin_border
            
            status = r['status']
            status_cell = ws.cell(row=row, column=9, value=status.upper())
            status_cell.border = self.thin_border
            
            # Color code
            bg = self.colors.get(status, 'FFFFFF')
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=bg)
            
            row += 1
        
        # Column widths
        widths = [6, 45, 8, 12, 12, 10, 40, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i)].width = w
    
    def _create_notfound_sheet(self, ws, items, results):
        """Sheet listing items that need manual pricing"""
        ws['A1'] = "ITEMS REQUIRING MANUAL PRICING"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['No.', 'Description', 'Quantity', 'Unit', 'Enter Price Here']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='FFC7CE')
            cell.border = self.thin_border
        
        row = 4
        count = 0
        for i, r in enumerate(results, 1):
            if r['status'] == 'notfound':
                count += 1
                ws.cell(row=row, column=1, value=count).border = self.thin_border
                ws.cell(row=row, column=2, value=r['description']).border = self.thin_border
                ws.cell(row=row, column=3, value=r['quantity']).border = self.thin_border
                ws.cell(row=row, column=4, value=r['unit']).border = self.thin_border
                price_cell = ws.cell(row=row, column=5, value='')
                price_cell.border = self.thin_border
                price_cell.fill = PatternFill('solid', fgColor='FFFF00')
                row += 1
        
        if count == 0:
            ws.cell(row=4, column=1, value="All items have prices! ✓")
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 18


def process_requisition(request_path, price_db_path, output_path=None):
    """Main processing function - handles PDF or Excel request files"""
    print("="*60)
    print("PHARMACY QUOTATION AUTOMATION v3")
    print("="*60)
    
    # Determine file type
    ext = request_path.lower().split('.')[-1]
    file_type = "Excel" if ext in ['xlsx', 'xls'] else "PDF"
    
    # Load prices
    print("\n[1/4] Loading price database...")
    price_db = PriceDatabase()
    loaded = ExcelLoader.load_prices(price_db_path, price_db)
    print(f"      ✓ {loaded} entries → {len(price_db.items)} unique items")
    
    # Parse request file
    print(f"\n[2/4] Parsing {file_type} request...")
    items = RequestParser.extract_items(request_path)
    print(f"      ✓ Found {len(items)} items")
    
    if items:
        print("\n      First 5 items:")
        for item in items[:5]:
            print(f"        • {item['description'][:45]}... ({item['quantity']} {item['unit']})")
    
    if not items:
        print("      ✗ No items found!")
        return None
    
    # Match
    print("\n[3/4] Matching prices...")
    found = sum(1 for item in items if price_db.find_price(item['description'])['found'])
    print(f"      ✓ {found}/{len(items)} items matched")
    
    # Generate
    print("\n[4/4] Generating quotation...")
    if not output_path:
        output_path = f"Quotation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    generator = QuotationGenerator()
    stats = generator.generate(items, price_db, output_path, 
                               title=os.path.basename(request_path))
    
    print(f"      ✓ Saved: {output_path}")
    print(f"\n      Results:")
    print(f"        ✓ Auto-priced:  {stats['found']}")
    print(f"        ⚠ Needs review: {stats['review']}")
    print(f"        ✗ Not found:    {stats['notfound']}")
    
    print("\n" + "="*60)
    print("COMPLETE!")
    print("="*60)
    
    return output_path


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 3:
        print("Usage: python pharmacy_automation_v3.py <request_file> <price_database.xlsx> [output.xlsx]")
        print("  request_file     — PDF or Excel requisition file")
        print("  price_database   — Excel file with historical pricing data")
        print("  output.xlsx      — (optional) output file name")
        sys.exit(1)

    request_path = sys.argv[1]
    price_db_path = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else None

    process_requisition(request_path, price_db_path, output_path)
